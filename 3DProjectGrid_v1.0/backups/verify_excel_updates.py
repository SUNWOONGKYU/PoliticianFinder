#!/usr/bin/env python3
"""
Excel 파일 업데이트 검증 스크립트
"""

import openpyxl

excel_file = r"G:\내 드라이브\Developement\PoliticianFinder\12D-GCDM_Grid\project_grid_v1.2_full_XY.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb['Project Grid']

print("=== 업데이트 검증 ===\n")

# P1B1 검증
print("P1B1 (Backend - FastAPI 초기화):")
print(f"  진도: {ws.cell(row=128, column=3).value}")
print(f"  상태: {ws.cell(row=129, column=3).value}")
print(f"  테스트/검토: {ws.cell(row=130, column=3).value}")

# P1D11 검증
print("\nP1D11 (Database - Alembic 초기화):")
print(f"  진도: {ws.cell(row=349, column=3).value}")
print(f"  상태: {ws.cell(row=350, column=3).value}")
print(f"  테스트/검토: {ws.cell(row=351, column=3).value}")

# P1A1 검증
print("\nP1A1 (AI/ML - Claude API 연동 준비):")
print(f"  진도: {ws.cell(row=602, column=3).value}")
print(f"  상태: {ws.cell(row=603, column=3).value}")
print(f"  테스트/검토: {ws.cell(row=604, column=3).value}")

print("\n" + "="*50)
print("검증 결과:")
print("="*50)

all_correct = True

# P1B1
if ws.cell(row=128, column=3).value == "100%":
    print("[OK] P1B1 진도: 100%")
else:
    print(f"[FAIL] P1B1 진도: {ws.cell(row=128, column=3).value}")
    all_correct = False

if ws.cell(row=129, column=3).value == "완료":
    print("[OK] P1B1 상태: 완료")
else:
    print(f"[FAIL] P1B1 상태: {ws.cell(row=129, column=3).value}")
    all_correct = False

if ws.cell(row=130, column=3).value == "통과":
    print("[OK] P1B1 테스트/검토: 통과")
else:
    print(f"[FAIL] P1B1 테스트/검토: {ws.cell(row=130, column=3).value}")
    all_correct = False

# P1D11
if ws.cell(row=349, column=3).value == "100%":
    print("[OK] P1D11 진도: 100%")
else:
    print(f"[FAIL] P1D11 진도: {ws.cell(row=349, column=3).value}")
    all_correct = False

if ws.cell(row=350, column=3).value == "완료":
    print("[OK] P1D11 상태: 완료")
else:
    print(f"[FAIL] P1D11 상태: {ws.cell(row=350, column=3).value}")
    all_correct = False

if ws.cell(row=351, column=3).value == "통과":
    print("[OK] P1D11 테스트/검토: 통과")
else:
    print(f"[FAIL] P1D11 테스트/검토: {ws.cell(row=351, column=3).value}")
    all_correct = False

# P1A1
if ws.cell(row=602, column=3).value == "100%":
    print("[OK] P1A1 진도: 100%")
else:
    print(f"[FAIL] P1A1 진도: {ws.cell(row=602, column=3).value}")
    all_correct = False

if ws.cell(row=603, column=3).value == "완료":
    print("[OK] P1A1 상태: 완료")
else:
    print(f"[FAIL] P1A1 상태: {ws.cell(row=603, column=3).value}")
    all_correct = False

if ws.cell(row=604, column=3).value == "통과":
    print("[OK] P1A1 테스트/검토: 통과")
else:
    print(f"[FAIL] P1A1 테스트/검토: {ws.cell(row=604, column=3).value}")
    all_correct = False

print("\n" + "="*50)
if all_correct:
    print("✓ 모든 업데이트가 성공적으로 적용되었습니다!")
else:
    print("✗ 일부 업데이트가 실패했습니다.")
print("="*50)
