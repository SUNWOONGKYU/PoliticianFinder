#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
양방향 동기화 테스트 스크립트
1단계: CSV 읽기 테스트
2단계: Excel 읽기 테스트
3단계: Excel → CSV 변환 테스트
"""

import sys
import csv
from pathlib import Path

if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl 라이브러리가 필요합니다.")
    sys.exit(1)


def test_csv_read(csv_path):
    """CSV 읽기 테스트"""
    print("\n" + "="*70)
    print("📄 Step 1: CSV 파일 읽기 테스트")
    print("="*70)

    try:
        csv_path = Path(csv_path)
        if not csv_path.exists():
            print(f"❌ 파일을 찾을 수 없습니다: {csv_path}")
            return False

        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

        print(f"✓ 파일 읽기 성공")
        print(f"✓ 총 행 수: {len(rows)}")
        print(f"✓ 첫 번째 행 (헤더): {rows[0][:3]}...")
        print(f"✓ 열 수: {len(rows[0])}")
        return rows

    except Exception as e:
        print(f"❌ CSV 읽기 실패: {e}")
        return None


def test_excel_read(excel_path):
    """Excel 읽기 테스트"""
    print("\n" + "="*70)
    print("📊 Step 2: Excel 파일 읽기 테스트")
    print("="*70)

    try:
        excel_path = Path(excel_path)
        if not excel_path.exists():
            print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
            return False

        wb = load_workbook(excel_path)
        print(f"✓ Excel 파일 열기 성공")
        print(f"✓ 시트 목록: {wb.sheetnames}")

        if "Project Grid" not in wb.sheetnames:
            print(f"❌ 'Project Grid' 시트를 찾을 수 없습니다.")
            return None

        ws = wb["Project Grid"]
        print(f"✓ 'Project Grid' 시트 선택")
        print(f"✓ 데이터 범위: {ws.dimensions}")

        # 샘플 데이터 읽기
        sample_rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            sample_rows.append(row)
            if idx >= 5:  # 처음 5개 행만
                break

        print(f"✓ 샘플 데이터 읽기 (첫 5행):")
        for idx, row in enumerate(sample_rows):
            print(f"   행 {idx+1}: {str(row[:3])+'...' if len(row) > 3 else row}")

        return ws

    except Exception as e:
        print(f"❌ Excel 읽기 실패: {e}")
        import traceback
        traceback.print_exc()
        return None


def test_excel_to_csv(excel_path, csv_backup_path):
    """Excel → CSV 변환 테스트"""
    print("\n" + "="*70)
    print("🔄 Step 3: Excel → CSV 변환 테스트")
    print("="*70)

    try:
        excel_path = Path(excel_path)
        csv_backup_path = Path(csv_backup_path)

        if not excel_path.exists():
            print(f"❌ Excel 파일을 찾을 수 없습니다: {excel_path}")
            return False

        print(f"✓ Excel 파일 읽기 중...")
        wb = load_workbook(excel_path)

        if "Project Grid" not in wb.sheetnames:
            print(f"❌ 'Project Grid' 시트를 찾을 수 없습니다.")
            return False

        ws = wb["Project Grid"]

        # 데이터 추출
        rows = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_data)
            row_count += 1

        print(f"✓ Excel에서 {row_count}개 행 추출")

        # CSV로 저장
        with open(csv_backup_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        print(f"✓ CSV 파일 생성: {csv_backup_path.name}")
        print(f"✓ 저장된 행 수: {len(rows)}")

        # 검증: 저장된 CSV 다시 읽기
        with open(csv_backup_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            verify_rows = list(reader)

        print(f"✓ 검증: 저장된 CSV 재확인 - {len(verify_rows)}개 행")

        if len(rows) == len(verify_rows):
            print(f"✅ Excel → CSV 변환 성공!")
            return True
        else:
            print(f"⚠️  행 수 불일치: {len(rows)} → {len(verify_rows)}")
            return False

    except Exception as e:
        print(f"❌ Excel → CSV 변환 실패: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """테스트 실행"""
    script_dir = Path(__file__).parent
    csv_file = script_dir / 'project_grid_v5.0_phase2d_complete.csv'
    excel_file = script_dir / 'project_grid_v5.0_phase2d_complete.xlsx'
    csv_backup_file = script_dir / 'project_grid_v5.0_phase2d_complete_TEST.csv'

    print("\n" + "🔍 양방향 동기화 테스트")
    print("="*70)

    # Step 1: CSV 읽기
    csv_data = test_csv_read(csv_file)
    if not csv_data:
        print("❌ 테스트 중단 (CSV 읽기 실패)")
        return False

    # Step 2: Excel 읽기
    excel_ws = test_excel_read(excel_file)
    if not excel_ws:
        print("❌ 테스트 중단 (Excel 읽기 실패)")
        return False

    # Step 3: Excel → CSV 변환
    success = test_excel_to_csv(excel_file, csv_backup_file)

    print("\n" + "="*70)
    if success:
        print("✅ 모든 테스트 성공!")
        print("\n다음 단계:")
        print("1. bidirectional_sync.py 실행하여 양방향 동기화 시작")
        print("2. Excel에서 셀 수정 후 저장")
        print("3. CSV 파일이 자동으로 업데이트되는지 확인")
        print("\n명령어: python bidirectional_sync.py")
    else:
        print("❌ 테스트 실패")

    print("="*70 + "\n")


if __name__ == '__main__':
    main()
