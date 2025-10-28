#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV ↔ Excel 양방향 동기화 시스템
CSV 파일이 변경되면 Excel 업데이트, Excel 파일이 변경되면 CSV 업데이트
"""

import os
import sys
import csv
import time
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# UTF-8 출력 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl 라이브러리가 설치되지 않았습니다.")
    print("   다음 명령으로 설치하세요: pip install openpyxl watchdog")
    sys.exit(1)


class BidirectionalSyncHandler(FileSystemEventHandler):
    """파일 변경 감지 및 동기화 핸들러"""

    def __init__(self, csv_path, excel_path):
        self.csv_path = Path(csv_path)
        self.excel_path = Path(excel_path)
        self.syncing = False
        self.last_sync_time = 0

    def on_modified(self, event):
        """파일 수정 감지"""
        if event.is_directory:
            return

        file_path = Path(event.src_path)

        # 동기화 중이면 무시 (무한 루프 방지)
        if self.syncing:
            return

        # 너무 빈번한 변경은 무시 (1초 이내)
        current_time = time.time()
        if current_time - self.last_sync_time < 1:
            return

        self.last_sync_time = current_time

        # CSV 파일 변경 감지
        if file_path.name == self.csv_path.name:
            self.syncing = True
            try:
                print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 📄 CSV 파일 변경 감지!")
                csv_to_excel(str(self.csv_path), str(self.excel_path))
                print("✅ Excel 파일 업데이트 완료\n")
            except Exception as e:
                print(f"❌ 동기화 실패: {e}\n")
            finally:
                self.syncing = False

        # Excel 파일 변경 감지
        elif file_path.name == self.excel_path.name:
            self.syncing = True
            try:
                print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 📊 Excel 파일 변경 감지!")
                excel_to_csv(str(self.excel_path), str(self.csv_path))
                print("✅ CSV 파일 업데이트 완료\n")
            except Exception as e:
                print(f"❌ 동기화 실패: {e}\n")
            finally:
                self.syncing = False


def csv_to_excel(csv_path, excel_path):
    """CSV를 Excel로 변환 (기존 함수 사용)"""
    from automation.csv_to_excel_with_colors import csv_to_excel_with_colors

    print(f"   CSV → Excel 변환 중...")
    csv_to_excel_with_colors(csv_path)


def excel_to_csv(excel_path, csv_path):
    """
    Excel 파일을 CSV로 변환 (역방향 동기화)

    Excel의 "Project Grid" 시트를 읽어서 CSV로 저장
    """
    try:
        print(f"   Excel → CSV 변환 중...")

        # Excel 파일 읽기
        wb = load_workbook(excel_path)

        # "Project Grid" 시트 선택
        if "Project Grid" not in wb.sheetnames:
            print(f"   ⚠️  'Project Grid' 시트를 찾을 수 없습니다.")
            return

        ws = wb["Project Grid"]

        # 수정 이력 기록을 위한 타임스탬프
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 기존 CSV 백업
        csv_path = Path(csv_path)
        if csv_path.exists():
            backup_path = csv_path.parent / f"{csv_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{csv_path.suffix}"
            csv_path.rename(backup_path)
            print(f"   📦 CSV 백업 생성: {backup_path.name}")

        # Excel 데이터 추출
        rows = []
        for row in ws.iter_rows(values_only=True):
            # None 값을 빈 문자열로 변환
            row_data = [str(cell) if cell is not None else "" for cell in row]
            rows.append(row_data)

        # CSV 파일에 쓰기
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        print(f"   ✓ CSV 파일 저장: {csv_path.name}")
        print(f"   ✓ 수정 일시: {timestamp}")
        print(f"   ✓ 추출 행 수: {len(rows)}")

    except Exception as e:
        print(f"   ❌ Excel → CSV 변환 실패: {e}")
        raise


def start_bidirectional_sync(csv_path, excel_path, watch_dir=None):
    """
    CSV ↔ Excel 양방향 동기화 시작

    Args:
        csv_path: CSV 파일 경로
        excel_path: Excel 파일 경로
        watch_dir: 감시할 디렉토리 (기본값: CSV 파일이 있는 디렉토리)
    """
    if watch_dir is None:
        watch_dir = str(Path(csv_path).parent)

    print("\n" + "="*70)
    print("🔄 CSV ↔ Excel 양방향 동기화 시작")
    print("="*70)
    print(f"📄 CSV 파일: {Path(csv_path).name}")
    print(f"📊 Excel 파일: {Path(excel_path).name}")
    print(f"📁 감시 디렉토리: {watch_dir}")
    print("\n기능:")
    print("  ✓ CSV 변경 → Excel 자동 업데이트")
    print("  ✓ Excel 변경 → CSV 자동 업데이트")
    print("  ✓ 수정 이력 자동 기록")
    print("  ✓ CSV 백업 자동 생성")
    print("\nCtrl+C를 눌러 종료하세요...")
    print("="*70 + "\n")

    # 파일 존재 여부 확인
    if not Path(csv_path).exists():
        print(f"❌ CSV 파일을 찾을 수 없습니다: {csv_path}")
        return False

    if not Path(excel_path).exists():
        print(f"⚠️  Excel 파일이 없습니다. 첫 변환을 수행합니다...")
        try:
            csv_to_excel(csv_path, excel_path)
            print("✅ Excel 파일 생성 완료\n")
        except Exception as e:
            print(f"❌ Excel 파일 생성 실패: {e}\n")
            return False

    # 파일 감지기 설정
    event_handler = BidirectionalSyncHandler(csv_path, excel_path)
    observer = Observer()
    observer.schedule(event_handler, watch_dir, recursive=False)

    try:
        observer.start()
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n" + "="*70)
        print("🛑 양방향 동기화 종료")
        print("="*70)
        observer.stop()

    observer.join()
    return True


def main():
    """메인 함수"""
    script_dir = Path(__file__).parent
    csv_file = script_dir / 'project_grid_v5.0_phase2d_complete.csv'
    excel_file = script_dir / 'project_grid_v5.0_phase2d_complete.xlsx'

    start_bidirectional_sync(str(csv_file), str(excel_file), str(script_dir))


if __name__ == '__main__':
    main()
