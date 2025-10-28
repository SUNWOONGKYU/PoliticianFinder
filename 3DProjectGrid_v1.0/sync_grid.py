"""
ProjectGrid CSV ↔ Excel 양방향 동기화 시스템

특징:
- CSV 파일 변경 감지 → Excel 자동 생성/업데이트
- Excel 파일 변경 감지 → CSV 자동 업데이트
- 상태에 따른 자동 셀 색상 적용
- Phase별 자동 그룹화
- 대시보드 자동 생성
- 실시간 동기화 또는 정기 동기화

사용법:
    python sync_grid.py --csv project_grid_v5.0.csv --xlsx project_grid_v5.0.xlsx --mode auto
"""

import csv
import os
import json
import sys
import time
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set
from dataclasses import dataclass
import argparse

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl needed. Install: pip install openpyxl")
    sys.exit(1)

# UTF-8 인코딩 설정
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')


@dataclass
class GridConfig:
    """프로젝트 그리드 설정"""
    # 색상 설정 (상태별)
    COLORS = {
        "완료": "90EE90",           # 초록색
        "진행": "FFFF99",           # 노란색
        "대기": "FFFFFF",           # 흰색
        "보류": "FFD700",           # 주황색
        "재작업필요": "FFB6C1"      # 빨간색
    }

    # 15개 필수 속성 (순서 중요)
    REQUIRED_ATTRIBUTES = [
        "작업ID",
        "업무",
        "작업지시서",
        "담당AI",
        "진도",
        "상태",
        "검증 방법",
        "테스트/검토",
        "자동화방식",
        "의존작업",
        "블로커",
        "비고",
        "수정 이력"
    ]

    # 헤더 스타일
    HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Area 헤더 스타일
    AREA_FILL = PatternFill(start_color="E7F0FF", end_color="E7F0FF", fill_type="solid")
    AREA_FONT = Font(bold=True, size=11)

    # 데이터 셀 스타일
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


class CSVGridParser:
    """CSV 파일 파서 - 3D 그리드 구조 복원"""

    def __init__(self, csv_path: str):
        self.csv_path = csv_path
        self.data = []
        self.headers = []
        self.areas = {}  # {area_name: {attribute: {phase: value}}}

    def parse(self) -> Dict:
        """CSV 파일 파싱 및 3D 그리드 구조 복원"""
        print(f"[CSV Parser] {self.csv_path} 파싱 중...")

        with open(self.csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            print("[ERROR] CSV 파일이 비어있습니다")
            return {}

        # 헤더 파싱
        self.headers = rows[0]
        print(f"  헤더: {self.headers[:3]}... ({len(self.headers)} 열)")

        # Phase 추출 (세 번째 열부터)
        phases = self.headers[2:] if len(self.headers) > 2 else []
        print(f"  Phase 수: {len(phases)}")

        # Area별 그룹화 및 데이터 추출
        current_area = None
        area_rows = {}

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(c == '' for c in row):
                continue

            area_name = row[0]
            attribute_name = row[1]

            # Area 헤더 행 (첫 열이 비어있지 않고, 두 번째 열이 비어있음)
            if area_name and not attribute_name:
                current_area = area_name
                area_rows[current_area] = {}
                print(f"  Area 발견: {current_area}")

            # 속성 행
            elif current_area and attribute_name:
                if current_area not in area_rows:
                    area_rows[current_area] = {}

                # 각 Phase의 값 추출
                values = row[2:] if len(row) > 2 else []
                area_rows[current_area][attribute_name] = {
                    phases[i]: values[i] if i < len(values) else ""
                    for i in range(len(phases))
                }

        self.areas = area_rows
        print(f"  총 Area: {len(area_rows)}")
        print(f"  총 속성 행: {sum(len(v) for v in area_rows.values())}")

        return {
            "headers": self.headers,
            "phases": phases,
            "areas": area_rows
        }

    def get_file_hash(self) -> str:
        """CSV 파일의 해시값 계산 (변경 감지용)"""
        with open(self.csv_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()


class ExcelGridGenerator:
    """Excel 파일 생성 및 포매팅"""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.wb = None
        self.ws = None
        self.config = GridConfig()

    def generate(self, parsed_data: Dict) -> None:
        """파싱된 데이터를 바탕으로 Excel 파일 생성"""
        print(f"[Excel Generator] {self.xlsx_path} 생성 중...")

        # 기존 파일 삭제
        if os.path.exists(self.xlsx_path):
            os.remove(self.xlsx_path)

        # 새 워크북 생성
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Project Grid"

        headers = parsed_data.get("headers", [])
        phases = parsed_data.get("phases", [])
        areas = parsed_data.get("areas", {})

        # 1. 헤더 행 생성
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.config.HEADER_FILL
            cell.font = self.config.HEADER_FONT
            cell.alignment = self.config.HEADER_ALIGNMENT

        self.ws.row_dimensions[1].height = 30

        # 2. 데이터 행 생성
        current_row = 2

        for area_name, attributes in areas.items():
            # Area 헤더 행
            area_cell = self.ws.cell(row=current_row, column=1, value=area_name)
            area_cell.fill = self.config.AREA_FILL
            area_cell.font = self.config.AREA_FONT
            self.ws.row_dimensions[current_row].height = 20
            current_row += 1

            # 속성 행들
            for attr_name, phase_values in attributes.items():
                attr_cell = self.ws.cell(row=current_row, column=2, value=attr_name)
                attr_cell.font = Font(bold=True)

                # Phase별 값 입력
                for col_idx, phase in enumerate(phases, 3):
                    value = phase_values.get(phase, "")
                    cell = self.ws.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = self.config.DATA_ALIGNMENT
                    cell.border = self.config.THIN_BORDER

                    # 상태 행에 색상 적용
                    if attr_name == "상태" and value:
                        self._apply_status_color(cell, value)

                self.ws.row_dimensions[current_row].height = 20
                current_row += 1

        # 3. 열 너비 조정
        self._adjust_column_widths(len(phases))

        # 4. 대시보드 시트 생성
        self._generate_dashboard(parsed_data)

        # 5. 파일 저장
        self.wb.save(self.xlsx_path)
        print(f"  ✓ Excel 파일 생성 완료: {self.xlsx_path}")
        print(f"  행 수: {current_row - 1}")
        print(f"  Area 수: {len(areas)}")

    def _apply_status_color(self, cell, status_value: str) -> None:
        """상태값에 따라 셀 색상 적용"""
        for status, color in self.config.COLORS.items():
            if status in status_value:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                break

    def _adjust_column_widths(self, phase_count: int) -> None:
        """열 너비 자동 조정"""
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 18
        for col_idx in range(3, 3 + phase_count):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = 25

    def _generate_dashboard(self, parsed_data: Dict) -> None:
        """대시보드 시트 생성"""
        print("  대시보드 시트 생성 중...")

        dashboard = self.wb.create_sheet("Dashboard")

        areas = parsed_data.get("areas", {})
        phases = parsed_data.get("phases", [])

        # 전체 진행률
        dashboard['A1'] = "전체 프로젝트 진행률"
        dashboard['A1'].font = Font(bold=True, size=14)

        # Phase별 진행률
        dashboard['A3'] = "Phase별 진행률"
        dashboard['A3'].font = Font(bold=True, size=12)

        row = 4
        dashboard['A4'] = "Phase"
        dashboard['B4'] = "완료"
        dashboard['C4'] = "진행"
        dashboard['D4'] = "대기"
        dashboard['E4'] = "진행률"

        for header_cell in ['A4', 'B4', 'C4', 'D4', 'E4']:
            dashboard[header_cell].font = Font(bold=True)
            dashboard[header_cell].fill = PatternFill(start_color="E7F0FF", fill_type="solid")

        row = 5
        for phase_idx, phase in enumerate(phases):
            dashboard[f'A{row}'] = phase
            row += 1

        # Area별 진행률
        dashboard['A20'] = "Area별 진행률"
        dashboard['A20'].font = Font(bold=True, size=12)

        row = 21
        dashboard['A21'] = "Area"
        dashboard['B21'] = "완료"
        dashboard['C21'] = "진행"
        dashboard['D21'] = "대기"
        dashboard['E21'] = "진행률"

        for header_cell in ['A21', 'B21', 'C21', 'D21', 'E21']:
            dashboard[header_cell].font = Font(bold=True)
            dashboard[header_cell].fill = PatternFill(start_color="E7F0FF", fill_type="solid")

        row = 22
        for area_name in areas.keys():
            dashboard[f'A{row}'] = area_name
            row += 1

        print(f"  ✓ 대시보드 생성 완료")

    def get_file_hash(self) -> str:
        """Excel 파일의 해시값 계산 (변경 감지용)"""
        if os.path.exists(self.xlsx_path):
            with open(self.xlsx_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        return ""


class ExcelGridReader:
    """Excel 파일에서 데이터 읽기"""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    def read(self) -> List[List]:
        """Excel 파일 데이터 읽기"""
        print(f"[Excel Reader] {self.xlsx_path} 읽기 중...")

        try:
            wb = openpyxl.load_workbook(self.xlsx_path)
            ws = wb.active

            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            print(f"  ✓ 읽기 완료: {len(data)} 행")
            return data
        except Exception as e:
            print(f"  [ERROR] Excel 읽기 실패: {e}")
            return []

    def get_file_hash(self) -> str:
        """Excel 파일의 해시값 계산"""
        if os.path.exists(self.xlsx_path):
            with open(self.xlsx_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        return ""


class CSVGridWriter:
    """CSV 파일 쓰기"""

    def __init__(self, csv_path: str):
        self.csv_path = csv_path

    def write(self, excel_data: List[List]) -> None:
        """Excel 데이터를 CSV로 저장"""
        print(f"[CSV Writer] {self.csv_path} 쓰기 중...")

        try:
            with open(self.csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in excel_data:
                    # None 값을 빈 문자열로 변환
                    clean_row = ['' if cell is None else str(cell) for cell in row]
                    writer.writerow(clean_row)

            print(f"  ✓ CSV 쓰기 완료: {len(excel_data)} 행")

            # 수정 이력 추가
            self._add_sync_log()
        except Exception as e:
            print(f"  [ERROR] CSV 쓰기 실패: {e}")

    def _add_sync_log(self) -> None:
        """동기화 로그 추가"""
        log_file = self.csv_path.replace('.csv', '_sync.log')
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().isoformat()}] Excel → CSV 동기화 완료\n")


class ProjectGridSync:
    """프로젝트 그리드 CSV ↔ Excel 양방향 동기화"""

    def __init__(self, csv_path: str, xlsx_path: str):
        self.csv_path = csv_path
        self.xlsx_path = xlsx_path

        self.csv_parser = CSVGridParser(csv_path)
        self.excel_gen = ExcelGridGenerator(xlsx_path)
        self.excel_reader = ExcelGridReader(xlsx_path)
        self.csv_writer = CSVGridWriter(csv_path)

        self.last_csv_hash = ""
        self.last_xlsx_hash = ""

    def sync_csv_to_xlsx(self) -> bool:
        """CSV → Excel 동기화 (정방향)"""
        print("\n" + "="*60)
        print("【CSV → Excel 동기화】")
        print("="*60)

        try:
            # CSV 파싱
            parsed = self.csv_parser.parse()
            if not parsed:
                print("[ERROR] CSV 파싱 실패")
                return False

            # Excel 생성
            self.excel_gen.generate(parsed)

            print("\n✓ 동기화 완료: CSV → Excel")
            return True
        except Exception as e:
            print(f"\n[ERROR] 동기화 실패: {e}")
            return False

    def sync_xlsx_to_csv(self) -> bool:
        """Excel → CSV 동기화 (역방향)"""
        print("\n" + "="*60)
        print("【Excel → CSV 동기화】")
        print("="*60)

        try:
            # Excel 읽기
            excel_data = self.excel_reader.read()
            if not excel_data:
                print("[ERROR] Excel 읽기 실패")
                return False

            # CSV 쓰기
            self.csv_writer.write(excel_data)

            print("\n✓ 동기화 완료: Excel → CSV")
            return True
        except Exception as e:
            print(f"\n[ERROR] 동기화 실패: {e}")
            return False

    def auto_sync(self, interval: int = 5) -> None:
        """자동 양방향 동기화 (파일 변경 감지)"""
        print(f"\n[Auto Sync] 시작 (감시 간격: {interval}초)")
        print("Ctrl+C를 누르면 종료합니다.\n")

        try:
            while True:
                csv_hash = self.csv_parser.get_file_hash()
                xlsx_hash = self.excel_reader.get_file_hash()

                # CSV 변경 감지
                if csv_hash != self.last_csv_hash:
                    print(f"\n[감지] CSV 파일 변경: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    self.sync_csv_to_xlsx()
                    self.last_csv_hash = csv_hash
                    self.last_xlsx_hash = self.excel_gen.get_file_hash()

                # Excel 변경 감지
                elif xlsx_hash != self.last_xlsx_hash:
                    print(f"\n[감지] Excel 파일 변경: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    self.sync_xlsx_to_csv()
                    self.last_xlsx_hash = xlsx_hash
                    self.last_csv_hash = self.csv_parser.get_file_hash()

                time.sleep(interval)
        except KeyboardInterrupt:
            print("\n\n[Auto Sync] 종료됨")

    def initial_sync(self) -> None:
        """초기 동기화 (CSV → Excel)"""
        print("\n[Initial Sync] 시작")
        self.sync_csv_to_xlsx()
        self.last_csv_hash = self.csv_parser.get_file_hash()
        self.last_xlsx_hash = self.excel_gen.get_file_hash()
        print("[Initial Sync] 완료\n")


def main():
    parser = argparse.ArgumentParser(
        description="프로젝트 그리드 CSV ↔ Excel 양방향 동기화"
    )
    parser.add_argument('--csv', required=True, help='CSV 파일 경로')
    parser.add_argument('--xlsx', required=True, help='Excel 파일 경로')
    parser.add_argument(
        '--mode',
        choices=['csv2xlsx', 'xlsx2csv', 'auto', 'init'],
        default='init',
        help='동기화 모드 (기본값: init)'
    )
    parser.add_argument('--interval', type=int, default=5, help='자동 동기화 간격(초)')

    args = parser.parse_args()

    # 동기화 객체 생성
    sync = ProjectGridSync(args.csv, args.xlsx)

    # 모드별 동기화 실행
    if args.mode == 'csv2xlsx':
        sync.sync_csv_to_xlsx()
    elif args.mode == 'xlsx2csv':
        sync.sync_xlsx_to_csv()
    elif args.mode == 'auto':
        sync.auto_sync(interval=args.interval)
    elif args.mode == 'init':
        sync.initial_sync()


if __name__ == '__main__':
    main()
