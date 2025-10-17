"""
CSV 파일을 Excel 파일로 자동 변환하는 스크립트
openpyxl을 사용하여 서식이 적용된 Excel 파일을 생성합니다.
"""

import csv
import os
import sys
from datetime import datetime
from typing import List

# UTF-8 출력 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl 라이브러리가 설치되지 않았습니다.")
    print("   다음 명령으로 설치하세요: pip install openpyxl")
    sys.exit(1)


def read_csv(filepath: str) -> List[List[str]]:
    """CSV 파일 읽기"""
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        return list(reader)


def create_excel_from_csv(csv_path: str, excel_path: str):
    """CSV 파일을 Excel 파일로 변환"""
    print("\n" + "="*70)
    print("CSV → Excel 변환")
    print("="*70)

    # CSV 읽기
    print(f"\n✓ CSV 파일 읽기: {csv_path}")
    data = read_csv(csv_path)

    # Excel 워크북 생성
    print("✓ Excel 워크북 생성")
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Grid"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    phase_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    phase_font = Font(bold=True, size=11)

    # 속성별 색상 정의
    task_id_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 작업ID, 업무: 노란색
    progress_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # 진도, 상태: 초록색
    test_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # 테스트/검토: 분홍색

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 데이터 쓰기
    print("✓ 데이터 Excel에 쓰는 중...")
    for row_idx, row_data in enumerate(data, 1):
        # 현재 행의 속성 확인 (2번째 열에 속성 이름이 있음)
        attribute = row_data[1] if len(row_data) > 1 else ""

        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border

            # 첫 번째 행 (Phase 헤더)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 영역 행 (Frontend, Backend, Database 등)
            elif col_idx == 1 and cell_value in ['Frontend', 'Backend (Supabase)', 'Database (Supabase)', 'RLS Policies', 'Authentication', 'Test & QA', 'DevOps & Infra', 'Security']:
                cell.fill = phase_fill
                cell.font = phase_font
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # 속성 열 (작업ID, 업무, 진도, 상태 등)
            elif col_idx == 2:
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # 속성별 색상 적용 (데이터 셀에만 적용, col_idx >= 3)
            elif col_idx >= 3:
                # 작업ID, 업무: 노란색
                if attribute in ['작업ID', '업무']:
                    cell.fill = task_id_fill
                    cell.alignment = center_alignment

                # 진도, 상태: 초록색
                elif attribute in ['진도', '상태']:
                    cell.fill = progress_fill
                    cell.alignment = center_alignment

                # 테스트/검토: 분홍색
                elif attribute == '테스트/검토':
                    cell.fill = test_fill
                    cell.alignment = center_alignment

    # 열 너비 조정
    print("✓ 열 너비 자동 조정")
    column_widths = {}
    for row in data:
        for idx, cell_value in enumerate(row):
            if cell_value:
                max_width = max(column_widths.get(idx, 0), len(str(cell_value)))
                column_widths[idx] = min(max_width, 50)  # 최대 50자

    for idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(idx + 1)].width = width + 2

    # 첫 두 열은 고정 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20

    # 첫 행 고정
    ws.freeze_panes = 'C2'

    # Excel 파일 저장
    wb.save(excel_path)
    print(f"✓ Excel 파일 저장: {excel_path}")

    # 파일 크기 출력
    file_size = os.path.getsize(excel_path) / 1024
    print(f"✓ 파일 크기: {file_size:.2f} KB")

    print("\n" + "="*70)
    print("✅ 변환 완료!")
    print("="*70)


def main():
    """메인 함수"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_file = 'project_grid_v2.0_supabase.csv'
    excel_file = 'project_grid_v2.0_supabase.xlsx'

    csv_path = os.path.join(script_dir, csv_file)
    excel_path = os.path.join(script_dir, excel_file)

    if not os.path.exists(csv_path):
        print(f"❌ CSV 파일을 찾을 수 없습니다: {csv_path}")
        return False

    try:
        create_excel_from_csv(csv_path, excel_path)
        return True
    except Exception as e:
        print(f"\n❌ 변환 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
