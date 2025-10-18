#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV to Excel Converter with Color Coding v1.0
CSV 파일을 읽어서 색상이 적용된 Excel 파일(.xlsx) 생성

Usage:
    python csv_to_excel_with_colors.py <csv_file>
"""

import sys
import csv
import json
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_colors():
    """colors.json 파일에서 색상 설정 로드"""
    colors_file = Path(__file__).parent / 'colors.json'

    # 기본 색상 (colors.json이 없을 때 사용)
    default_colors = {
        'phase1': 'BBDEFB', 'phase2': 'C8E6C9', 'phase3': 'FFF9C4',
        'phase4': 'FFE0B2', 'phase5': 'FFCDD2',
        'Frontend': 'E3F2FD', 'Backend': 'F1F8E9', 'Database': 'FFF8E1',
        'Test': 'F3E5F5', 'DevOps': 'ECEFF1', 'AI/ML': 'FCE4EC',
        '0%': 'FFFFFF', '25%': 'E0E0E0', '50%': 'BDBDBD',
        '75%': '757575', '100%': '4CAF50',
        '대기': 'FFFFFF', '진행중': 'E3F2FD', '검토중': 'FFF3E0',
        '완료': 'E8F5E9', '보류': 'FFEBEE',
        '의존없음': 'FFFFFF', '의존있음': 'FFF3E0',
        '없음': 'FFFFFF', '의존성 대기': 'FFF3E0', '기술 이슈': 'FFEBEE',
        '요구사항 불명확': 'FFF9C4', '외부 의존': 'E1F5FE',
        'header': 'B0BEC5'
    }

    if not colors_file.exists():
        print(f"[WARNING] colors.json 파일을 찾을 수 없습니다. 기본 색상을 사용합니다.")
        return default_colors

    try:
        with open(colors_file, 'r', encoding='utf-8') as f:
            color_config = json.load(f)

        # JSON 구조를 평면화
        colors = {}
        colors.update(color_config.get('phase', {}))
        colors.update(color_config.get('area', {}))
        colors.update(color_config.get('progress', {}))
        colors.update(color_config.get('state', {}))
        colors.update(color_config.get('blocker', {}))
        colors.update(color_config.get('other', {}))

        # 의존작업 처리
        dep = color_config.get('dependency', {})
        colors['의존없음'] = dep.get('없음', 'FFFFFF')
        colors['의존있음'] = dep.get('있음', 'FFF3E0')

        # 작업 행 색상
        task = color_config.get('task', {})
        colors['작업ID_배경'] = task.get('작업ID', 'E1F5FE')
        colors['업무_배경'] = task.get('업무', 'FFF9C4')

        print(f"[OK] colors.json 파일 로드 완료")
        return colors

    except Exception as e:
        print(f"[ERROR] colors.json 로드 실패: {e}. 기본 색상을 사용합니다.")
        return default_colors

# 색상 로드
COLORS = load_colors()

def get_phase_color(cell_value):
    """Phase 컬럼의 색상 반환"""
    if 'Phase 1' in str(cell_value):
        return COLORS['phase1']
    elif 'Phase 2' in str(cell_value):
        return COLORS['phase2']
    elif 'Phase 3' in str(cell_value):
        return COLORS['phase3']
    elif 'Phase 4' in str(cell_value):
        return COLORS['phase4']
    elif 'Phase 5' in str(cell_value):
        return COLORS['phase5']
    return None

def get_area_color(cell_value):
    """영역 컬럼의 색상 반환"""
    return COLORS.get(str(cell_value).strip(), None)

def get_progress_color(cell_value):
    """진도율의 색상 반환"""
    return COLORS.get(str(cell_value).strip(), None)

def get_state_color(cell_value):
    """상태의 색상 반환"""
    return COLORS.get(str(cell_value).strip(), None)

def get_dependency_color(cell_value):
    """의존작업의 색상 반환"""
    val = str(cell_value).strip()
    if val == '없음' or val == '':
        return COLORS.get('의존없음', None)
    else:
        return COLORS.get('의존있음', None)

def get_blocker_color(cell_value):
    """블로커의 색상 반환"""
    return COLORS.get(str(cell_value).strip(), None)

def apply_cell_style(cell, bg_color=None, bold=False, font_size=11):
    """셀에 스타일 적용"""
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    cell.font = Font(name='Arial', size=font_size, bold=bold)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 테두리
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def analyze_grid_data(rows):
    """그리드 데이터를 분석하여 통계 반환"""
    stats = {
        'total_tasks': 0,
        'completed_tasks': 0,
        'in_progress_tasks': 0,
        'reviewing_tasks': 0,
        'blocked_tasks': 0,
        'phase_stats': {},
        'area_stats': {},
        'blocker_count': 0,
        'dependency_count': 0
    }

    current_area = None

    for row_idx, row in enumerate(rows):
        if row_idx == 0:  # 헤더 행 스킵
            continue

        # 영역 행 감지
        if row[0] and row[0] in ['Frontend', 'Backend', 'Database', 'Test', 'DevOps', 'AI/ML']:
            current_area = row[0]
            if current_area not in stats['area_stats']:
                stats['area_stats'][current_area] = {'total': 0, 'completed': 0}
            continue

        # 속성 행 확인
        if len(row) > 1 and row[1]:
            attr = row[1]

            # 작업ID 행 = 새로운 작업
            if attr == '작업ID':
                stats['total_tasks'] += 1
                if current_area:
                    stats['area_stats'][current_area]['total'] += 1

            # 상태 행
            elif attr == '상태':
                for phase_idx, value in enumerate(row[2:], start=1):
                    if value:
                        phase_name = f'Phase {phase_idx}'
                        if phase_name not in stats['phase_stats']:
                            stats['phase_stats'][phase_name] = {'total': 0, 'completed': 0}

                        stats['phase_stats'][phase_name]['total'] += 1

                        if value == '완료':
                            stats['completed_tasks'] += 1
                            stats['phase_stats'][phase_name]['completed'] += 1
                            if current_area:
                                stats['area_stats'][current_area]['completed'] += 1
                        elif value == '진행중':
                            stats['in_progress_tasks'] += 1
                        elif value == '검토중':
                            stats['reviewing_tasks'] += 1
                        elif value == '보류':
                            stats['blocked_tasks'] += 1

            # 블로커 행
            elif attr == '블로커':
                for value in row[2:]:
                    if value and value != '없음':
                        stats['blocker_count'] += 1

            # 의존작업 행
            elif attr == '의존작업':
                for value in row[2:]:
                    if value and value != '없음':
                        stats['dependency_count'] += 1

    return stats

def create_dashboard_sheet(wb, stats):
    """대시보드 시트 생성"""
    ws = wb.create_sheet("대시보드", 0)  # 첫 번째 시트로 추가

    # 제목
    ws['A1'] = '📊 프로젝트 진행 현황 대시보드'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws.merge_cells('A1:D1')

    row = 3

    # 전체 진행률
    ws[f'A{row}'] = '전체 진행률'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    total_progress = (stats['completed_tasks'] / stats['total_tasks'] * 100) if stats['total_tasks'] > 0 else 0
    ws[f'B{row}'] = f"{total_progress:.1f}%"
    ws[f'B{row}'].font = Font(size=12)
    ws[f'C{row}'] = f"({stats['completed_tasks']}/{stats['total_tasks']} 작업)"
    ws[f'B{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    row += 2

    # 작업 상태별 현황
    ws[f'A{row}'] = '작업 상태'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = '완료'
    ws[f'B{row}'] = stats['completed_tasks']
    ws[f'B{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    row += 1

    ws[f'A{row}'] = '진행중'
    ws[f'B{row}'] = stats['in_progress_tasks']
    ws[f'B{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    row += 1

    ws[f'A{row}'] = '검토중'
    ws[f'B{row}'] = stats['reviewing_tasks']
    ws[f'B{row}'].fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    row += 1

    ws[f'A{row}'] = '보류'
    ws[f'B{row}'] = stats['blocked_tasks']
    ws[f'B{row}'].fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    row += 1

    ws[f'A{row}'] = '대기'
    waiting = stats['total_tasks'] - stats['completed_tasks'] - stats['in_progress_tasks'] - stats['reviewing_tasks'] - stats['blocked_tasks']
    ws[f'B{row}'] = waiting

    row += 2

    # Phase별 진행률
    ws[f'A{row}'] = 'Phase별 진행률'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for phase_name in sorted(stats['phase_stats'].keys()):
        phase = stats['phase_stats'][phase_name]
        progress = (phase['completed'] / phase['total'] * 100) if phase['total'] > 0 else 0
        ws[f'A{row}'] = phase_name
        ws[f'B{row}'] = f"{progress:.1f}%"
        ws[f'C{row}'] = f"({phase['completed']}/{phase['total']})"

        if progress >= 75:
            ws[f'B{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        elif progress >= 25:
            ws[f'B{row}'].fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        else:
            ws[f'B{row}'].fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

        row += 1

    row += 1

    # 영역별 진행률
    ws[f'A{row}'] = '영역별 진행률'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for area_name in ['Frontend', 'Backend', 'Database', 'Test', 'DevOps', 'AI/ML']:
        if area_name in stats['area_stats']:
            area = stats['area_stats'][area_name]
            progress = (area['completed'] / area['total'] * 100) if area['total'] > 0 else 0
            ws[f'A{row}'] = area_name
            ws[f'B{row}'] = f"{progress:.1f}%"
            ws[f'C{row}'] = f"({area['completed']}/{area['total']})"

            if progress >= 75:
                ws[f'B{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            elif progress >= 25:
                ws[f'B{row}'].fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            else:
                ws[f'B{row}'].fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

            row += 1

    row += 1

    # 블로커 및 의존성
    ws[f'A{row}'] = '⚠️ 주의사항'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = '블로커 있는 작업'
    ws[f'B{row}'] = stats['blocker_count']
    if stats['blocker_count'] > 0:
        ws[f'B{row}'].fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    row += 1

    ws[f'A{row}'] = '의존작업 있는 작업'
    ws[f'B{row}'] = stats['dependency_count']
    ws[f'B{row}'].fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

def backup_csv(csv_path):
    """CSV 파일을 백업하고 오래된 백업 정리"""
    csv_path = Path(csv_path)

    # 백업 폴더 경로
    backup_dir = csv_path.parent / 'backups'
    backup_dir.mkdir(exist_ok=True)

    # 타임스탬프 생성
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 백업 파일명
    backup_filename = f"{csv_path.stem}_backup_{timestamp}{csv_path.suffix}"
    backup_path = backup_dir / backup_filename

    # 백업 수행
    shutil.copy2(csv_path, backup_path)
    print(f"[OK] CSV 백업 완료: {backup_path.name}")

    # 오래된 백업 정리 (최근 10개만 유지)
    backups = sorted(backup_dir.glob(f"{csv_path.stem}_backup_*{csv_path.suffix}"),
                    key=lambda p: p.stat().st_mtime,
                    reverse=True)

    if len(backups) > 10:
        for old_backup in backups[10:]:
            old_backup.unlink()
            print(f"[OK] 오래된 백업 삭제: {old_backup.name}")

    return backup_path

def auto_update_blockers(rows):
    """
    의존작업과 상태를 분석하여 블로커를 자동 업데이트

    로직:
    - 의존작업이 있고, 해당 작업이 '완료' 상태가 아니면 → 블로커에 '의존성 대기' 설정
    - 의존작업이 모두 완료되었으면 → 블로커 '없음'으로 변경
    """
    # 작업 상태 맵 구축 (작업ID -> 상태)
    task_status_map = {}

    current_area = None
    task_id_row = None
    status_row = None

    for row_idx, row in enumerate(rows):
        # 영역 행 감지
        if row[0] and row[0] in ['Frontend', 'Backend', 'Database', 'Test', 'DevOps', 'AI/ML']:
            current_area = row[0]
            continue

        # 속성 행 확인
        if len(row) > 1 and row[1]:
            attr = row[1]

            # 작업ID 행 저장
            if attr == '작업ID':
                task_id_row = row_idx
                for col_idx, task_id in enumerate(row[2:], start=2):
                    if task_id and task_id.strip():
                        task_status_map[task_id.strip()] = {
                            'status': '대기',
                            'row_idx': task_id_row,
                            'col_idx': col_idx
                        }

            # 상태 행으로 상태 맵 업데이트
            elif attr == '상태':
                status_row = row_idx
                for col_idx, status in enumerate(row[2:], start=2):
                    if status and task_id_row:
                        task_id = rows[task_id_row][col_idx]
                        if task_id and task_id.strip() in task_status_map:
                            task_status_map[task_id.strip()]['status'] = status.strip()

    # 블로커 자동 업데이트
    updated_count = 0

    for row_idx, row in enumerate(rows):
        if len(row) > 1 and row[1]:
            attr = row[1]

            # 의존작업 행 찾기
            if attr == '의존작업':
                dependency_row_idx = row_idx

                # 블로커 행은 의존작업 행 바로 다음
                blocker_row_idx = dependency_row_idx + 1

                # 블로커 행이 존재하는지 확인
                if blocker_row_idx < len(rows) and len(rows[blocker_row_idx]) > 1:
                    if rows[blocker_row_idx][1] == '블로커':

                        # 각 작업별로 의존성 체크
                        for col_idx in range(2, len(row)):
                            dependency_value = row[col_idx].strip() if row[col_idx] else ''

                            # 의존작업이 없으면 블로커도 '없음'
                            if not dependency_value or dependency_value == '없음':
                                if rows[blocker_row_idx][col_idx] == '의존성 대기':
                                    rows[blocker_row_idx][col_idx] = '없음'
                                    updated_count += 1
                                continue

                            # 의존작업이 있는 경우
                            dependencies = [d.strip() for d in dependency_value.split(',')]
                            all_completed = True

                            for dep_task_id in dependencies:
                                if dep_task_id in task_status_map:
                                    dep_status = task_status_map[dep_task_id]['status']
                                    if dep_status != '완료':
                                        all_completed = False
                                        break
                                else:
                                    # 의존 작업을 찾을 수 없는 경우 (외부 의존일 수 있음)
                                    all_completed = False

                            # 블로커 자동 설정
                            current_blocker = rows[blocker_row_idx][col_idx].strip() if rows[blocker_row_idx][col_idx] else ''

                            if not all_completed:
                                # 의존작업이 완료되지 않았으면 '의존성 대기'
                                if current_blocker != '의존성 대기' and current_blocker != '기술 이슈' and current_blocker != '요구사항 불명확' and current_blocker != '외부 의존':
                                    # 다른 블로커가 없으면 의존성 대기로 설정
                                    if current_blocker == '없음' or not current_blocker:
                                        rows[blocker_row_idx][col_idx] = '의존성 대기'
                                        updated_count += 1
                            else:
                                # 의존작업이 모두 완료되었으면 의존성 대기 해제
                                if current_blocker == '의존성 대기':
                                    rows[blocker_row_idx][col_idx] = '없음'
                                    updated_count += 1

    if updated_count > 0:
        print(f"[OK] 블로커 자동 업데이트: {updated_count}개 작업")

    return rows

def csv_to_excel_with_colors(csv_path):
    """CSV 파일을 색상이 적용된 Excel 파일로 변환"""

    csv_path = Path(csv_path)
    if not csv_path.exists():
        print(f"[ERROR] 파일을 찾을 수 없습니다: {csv_path}")
        return

    # CSV 백업
    backup_csv(csv_path)

    # Excel 파일 경로 (같은 이름, 확장자만 .xlsx)
    excel_path = csv_path.with_suffix('.xlsx')

    # CSV 읽기
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # 블로커 자동 업데이트 (의존작업 기반)
    rows = auto_update_blockers(rows)

    # 하이퍼링크 통계 초기화
    hyperlink_stats = {'total': 0, 'exists': 0, 'missing': 0}

    # Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Grid"

    # 통계 분석
    stats = analyze_grid_data(rows)

    # 데이터 쓰기 및 스타일 적용
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 헤더 행 (첫 번째 행)
            if row_idx == 1:
                bg_color = get_phase_color(value) if col_idx > 1 else COLORS['header']
                apply_cell_style(cell, bg_color=bg_color, bold=True, font_size=12)

            # 영역 컬럼 (첫 번째 컬럼, Frontend, Backend 등)
            elif col_idx == 1 and value.strip():
                bg_color = get_area_color(value)
                apply_cell_style(cell, bg_color=bg_color, bold=True, font_size=11)

            # 데이터 행
            else:
                # 행 패턴 분석 (각 작업은 10개 행으로 구성)
                # 1: 작업ID, 2: 업무, 3: 작업지시서, 4: 담당AI
                # 5: 진도, 6: 완료, 7: 테스트/검토, 8: 자동화방식
                # 9: 블로커, 10: 구분선

                # 작업 시작 행 찾기 (영역 행 이후)
                area_row = row_idx
                while area_row > 1 and not ws.cell(row=area_row, column=1).value:
                    area_row -= 1

                offset = row_idx - area_row
                line_type = offset % 10  # 0~9 반복

                bg_color = None
                # 속성 컬럼(2번째 컬럼) 값으로 색상 결정
                attr_value = ws.cell(row=row_idx, column=2).value

                # 작업ID 행
                if attr_value == '작업ID':
                    bg_color = COLORS.get('작업ID_배경', 'E1F5FE')

                # 업무 행
                elif attr_value == '업무':
                    bg_color = COLORS.get('업무_배경', 'FFF9C4')

                # 작업지시서 행 - 하이퍼링크 추가 ⭐ NEW
                elif attr_value == '작업지시서':
                    bg_color = COLORS.get('작업지시서_배경', 'E8F5E9')  # 연한 초록색

                    # 배경색과 테두리만 적용 (폰트는 나중에)
                    if bg_color:
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # 작업지시서 파일 경로가 있으면 하이퍼링크 추가
                    if value and value.strip() and value.strip().startswith('tasks/'):
                        hyperlink_stats['total'] += 1
                        file_path = value.strip()
                        # 절대 경로 생성 (Google Drive 경로)
                        abs_path = (Path(csv_path).parent / file_path).resolve()

                        # 파일이 실제로 존재하는지 확인
                        if abs_path.exists():
                            hyperlink_stats['exists'] += 1
                            # file:// URL로 변환
                            file_url = abs_path.as_uri()
                            cell.hyperlink = file_url
                            # 하이퍼링크 스타일 (파란색, 밑줄)
                            cell.font = Font(name='Arial', color="0563C1", underline="single", size=10)
                        else:
                            hyperlink_stats['missing'] += 1
                            # 파일이 없으면 주황색으로 표시 (작성 필요)
                            cell.font = Font(name='Arial', color="FF6600", size=10)
                    else:
                        # 파일 경로가 없으면 일반 폰트
                        cell.font = Font(name='Arial', size=10)

                    # 작업지시서는 apply_cell_style 건너뛰기 (이미 처리했으므로)
                    continue

                # 진도 행
                elif attr_value == '진도':
                    bg_color = get_progress_color(value)

                # 상태 행
                elif attr_value == '상태':
                    bg_color = get_state_color(value)

                # 의존작업 행
                elif attr_value == '의존작업':
                    bg_color = get_dependency_color(value)

                # 블로커 행
                elif attr_value == '블로커':
                    bg_color = get_blocker_color(value)

                apply_cell_style(cell, bg_color=bg_color, bold=False, font_size=10)

    # 열 너비 자동 조정
    for col_idx in range(1, len(rows[0]) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 15  # 영역 컬럼
        else:
            ws.column_dimensions[col_letter].width = 35  # Phase 컬럼들

    # 행 높이 조정
    for row_idx in range(1, len(rows) + 1):
        ws.row_dimensions[row_idx].height = 18

    # 첫 행과 첫 열 고정
    ws.freeze_panes = 'B2'

    # 대시보드 시트 생성 (첫 번째 시트로)
    create_dashboard_sheet(wb, stats)

    # 저장
    wb.save(excel_path)
    print(f"[OK] Excel 파일 생성 완료: {excel_path}")
    print(f"   - 총 {len(rows)}개 행")
    print(f"   - 시트: 대시보드 (진행률 요약) + Project Grid (상세 그리드)")
    print(f"   - 전체 진행률: {stats['completed_tasks']}/{stats['total_tasks']} 작업 완료")
    print(f"   - 색상 적용: Phase별, 영역별, 진도율, 상태, 의존작업, 블로커")
    print(f"   - 하이퍼링크: {hyperlink_stats['exists']}개 연결됨, {hyperlink_stats['missing']}개 미작성 (총 {hyperlink_stats['total']}개)")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("사용법: python csv_to_excel_with_colors.py <csv_file>")
        print("예제: python csv_to_excel_with_colors.py project_grid_v3.1_XY.csv")
        sys.exit(1)

    csv_file = sys.argv[1]
    csv_to_excel_with_colors(csv_file)
