#!/usr/bin/env python3
"""
Excel 파일에서 완료된 작업의 진도를 업데이트하는 스크립트
"""

import openpyxl
import shutil
from pathlib import Path

# Excel 파일 경로
excel_file = r"G:\내 드라이브\Developement\PoliticianFinder\12D-GCDM_Grid\project_grid_v1.2_full_XY.xlsx"

# 완료된 작업 정보
completed_tasks = {
    'P1A1': {
        '진도': '100%',
        '상태': '완료',
        '테스트/검토': '통과'
    },
    'P1B1': {
        '진도': '100%',
        '상태': '완료',
        '테스트/검토': '통과'
    },
    'P1D11': {
        '진도': '100%',
        '상태': '완료',
        '테스트/검토': '통과'
    }
}

def update_excel():
    """Excel 파일 업데이트"""

    # 백업 생성
    backup_file = excel_file + '.bak'
    shutil.copy2(excel_file, backup_file)
    print(f"[OK] Backup created: {backup_file}")

    # Excel 파일 열기
    wb = openpyxl.load_workbook(excel_file)

    # 모든 시트 확인
    updated_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[INFO] Checking sheet: {sheet_name}")

        # 작업ID 행 찾기
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value

            if cell_value == '작업ID':
                print(f"[INFO] Found '작업ID' at row {row_idx}")

                # 작업ID 행에서 각 작업 ID 찾기
                work_ids = {}
                for col_idx in range(3, ws.max_column + 1):
                    work_id = ws.cell(row=row_idx, column=col_idx).value
                    if work_id and work_id in completed_tasks:
                        work_ids[work_id] = col_idx
                        print(f"[INFO] Found {work_id} at column {col_idx}")

                if not work_ids:
                    continue

                # 진도, 상태, 테스트/검토 행 찾기 (작업ID 행 아래 4, 5, 6번째)
                progress_row = row_idx + 4
                status_row = row_idx + 5
                test_row = row_idx + 6

                # 진도 행 확인
                if ws.cell(row=progress_row, column=2).value == '진도':
                    for work_id, col_idx in work_ids.items():
                        old_value = ws.cell(row=progress_row, column=col_idx).value
                        new_value = completed_tasks[work_id]['진도']
                        ws.cell(row=progress_row, column=col_idx, value=new_value)
                        print(f"[UPDATE] {work_id} 진도: {old_value} -> {new_value}")
                        updated_count += 1

                # 상태 행 확인
                if ws.cell(row=status_row, column=2).value == '상태':
                    for work_id, col_idx in work_ids.items():
                        old_value = ws.cell(row=status_row, column=col_idx).value
                        new_value = completed_tasks[work_id]['상태']
                        ws.cell(row=status_row, column=col_idx, value=new_value)
                        print(f"[UPDATE] {work_id} 상태: {old_value} -> {new_value}")
                        updated_count += 1

                # 테스트/검토 행 확인
                if ws.cell(row=test_row, column=2).value == '테스트/검토':
                    for work_id, col_idx in work_ids.items():
                        old_value = ws.cell(row=test_row, column=col_idx).value
                        new_value = completed_tasks[work_id]['테스트/검토']
                        ws.cell(row=test_row, column=col_idx, value=new_value)
                        print(f"[UPDATE] {work_id} 테스트/검토: {old_value} -> {new_value}")
                        updated_count += 1

    # Excel 파일 저장
    wb.save(excel_file)
    print(f"\n[OK] Excel file updated successfully!")
    print(f"[OK] Total updates: {updated_count}")
    print(f"[OK] Updated file: {excel_file}")
    print(f"[OK] Backup file: {backup_file}")

if __name__ == '__main__':
    try:
        update_excel()
    except Exception as e:
        print(f"[ERROR] Failed to update Excel: {str(e)}")
        import traceback
        traceback.print_exc()
