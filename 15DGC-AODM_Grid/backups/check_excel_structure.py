#!/usr/bin/env python3
"""
Excel 파일 구조 확인 스크립트
"""

import openpyxl

excel_file = r"G:\내 드라이브\Developement\PoliticianFinder\12D-GCDM_Grid\project_grid_v1.2_full_XY.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb['Project Grid']

# P1B1 주변 확인 (row 124)
print("=== P1B1 (row 124) ===")
for row_idx in range(124, 135):
    col2 = ws.cell(row=row_idx, column=2).value
    col3 = ws.cell(row=row_idx, column=3).value
    print(f"Row {row_idx}: Col2={col2}, Col3={col3}")

print("\n=== P1D11 (row 345) ===")
for row_idx in range(345, 356):
    col2 = ws.cell(row=row_idx, column=2).value
    col3 = ws.cell(row=row_idx, column=3).value
    print(f"Row {row_idx}: Col2={col2}, Col3={col3}")

print("\n=== P1A1 (row 598) ===")
for row_idx in range(598, 609):
    col2 = ws.cell(row=row_idx, column=2).value
    col3 = ws.cell(row=row_idx, column=3).value
    print(f"Row {row_idx}: Col2={col2}, Col3={col3}")
