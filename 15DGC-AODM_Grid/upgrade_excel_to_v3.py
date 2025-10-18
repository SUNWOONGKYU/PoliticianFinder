#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
15DGC-AODM v3.0 엑셀 업그레이드 스크립트
project_grid_v2.0_supabase.xlsx를 v3.0으로 업그레이드
- "검증 방법" 행 추가 (상태 다음, 테스트/검토 이전)
- 파일명을 v3.0으로 변경
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import sys
import os

def upgrade_excel_to_v3(input_file, output_file):
    """엑셀 파일을 v3.0으로 업그레이드"""

    print(f"[+] {input_file} 파일을 읽는 중...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 스타일 정의 (기존 "상태" 행과 유사하게)
    verification_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")  # 연한 시안
    header_font = Font(name='맑은 고딕', size=9, bold=True)
    cell_font = Font(name='맑은 고딕', size=9)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 모든 행을 역순으로 순회하면서 "상태" 행 찾기
    max_row = ws.max_row
    rows_to_insert = []

    for row_idx in range(max_row, 0, -1):
        cell_b = ws.cell(row=row_idx, column=2)

        # "상태" 행을 찾으면
        if cell_b.value == '상태':
            # 다음 행이 "테스트/검토"인지 확인
            next_row_idx = row_idx + 1
            if next_row_idx <= max_row:
                next_cell_b = ws.cell(row=next_row_idx, column=2)
                if next_cell_b.value == '테스트/검토':
                    rows_to_insert.append(next_row_idx)

    print(f"[*] {len(rows_to_insert)}개의 작업 블록을 발견했습니다.")

    # 역순으로 정렬 (큰 행 번호부터 처리해야 행 번호가 밀리지 않음)
    rows_to_insert.sort(reverse=True)

    # 각 위치에 "검증 방법" 행 삽입
    for insert_row_idx in rows_to_insert:
        # 새 행 삽입
        ws.insert_rows(insert_row_idx)

        # B열에 "검증 방법" 입력
        verification_cell = ws.cell(row=insert_row_idx, column=2)
        verification_cell.value = '검증 방법'
        verification_cell.fill = verification_fill
        verification_cell.font = header_font
        verification_cell.alignment = center_alignment
        verification_cell.border = thin_border

        # 같은 행의 "상태" 행에서 열 개수 확인
        state_row_idx = insert_row_idx - 1
        max_col = ws.max_column

        # C열부터 끝까지 기본값 설정
        for col_idx in range(3, max_col + 1):
            state_cell = ws.cell(row=state_row_idx, column=col_idx)
            new_cell = ws.cell(row=insert_row_idx, column=col_idx)

            # 상태 셀에 값이 있고, 대기/빈값이 아니면 "Build Test" 입력
            if state_cell.value and state_cell.value not in ['대기', '-', '']:
                new_cell.value = 'Build Test'
            else:
                new_cell.value = '-'

            # 스타일 적용
            new_cell.fill = verification_fill
            new_cell.font = cell_font
            new_cell.alignment = center_alignment
            new_cell.border = thin_border

    print(f"[+] {len(rows_to_insert)}개 작업 블록에 '검증 방법' 행이 추가되었습니다.")

    # 파일 저장
    print(f"[+] {output_file} 파일을 저장하는 중...")
    wb.save(output_file)
    print(f"[+] 엑셀 파일 업그레이드 완료!")
    print(f"    총 {ws.max_row}행, {ws.max_column}열")

if __name__ == '__main__':
    input_file = 'project_grid_v2.0_supabase.xlsx'
    output_file = 'project_grid_v3.0_supabase.xlsx'

    if not os.path.exists(input_file):
        print(f"[!] 오류: {input_file} 파일을 찾을 수 없습니다.")
        sys.exit(1)

    upgrade_excel_to_v3(input_file, output_file)
